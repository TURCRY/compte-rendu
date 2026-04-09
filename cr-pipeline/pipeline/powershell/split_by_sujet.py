#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
split_by_sujet.py
Découpe out/global.json (agrégation par sujets) en N fichiers JSON (1 par sujet),
en s'appuyant sur Sujets.xlsx (colonne Numero + Titre).
Optionnel : filtre aussi debrief.json par sujet.

Usage:
  python split_by_sujet.py --global out/global.json --sujets Sujets.xlsx --out out/sujets
  python split_by_sujet.py --global out/global.json --sujets Sujets.xlsx --out out/sujets --debrief out/debrief.json
"""

from __future__ import annotations

import argparse
import json
import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


@dataclass(frozen=True)
class SujetRef:
    numero: int
    titre: str


def _read_json(path: str) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _write_json(path: str, obj: Any) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def _ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _norm_str(x: Any) -> str:
    return str(x).strip() if x is not None else ""


def _safe_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        try:
            return int(x)
        except Exception:
            return None
    s = _norm_str(x)
    if not s:
        return None
    # tolère "01", "1.0", etc.
    try:
        if "." in s:
            return int(float(s))
        return int(s)
    except Exception:
        return None


def _load_sujets_xlsx(path: str) -> List[SujetRef]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Trouver la ligne d'en-tête (on suppose première ligne)
    headers = {}
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            for idx, val in enumerate(row or ()):
                key = _norm_str(val).lower()
                if key:
                    headers[key] = idx
            break

    # Colonnes attendues : Numero, Titre (insensible à la casse)
    idx_num = headers.get("numero")
    idx_titre = headers.get("titre")

    if idx_num is None:
        raise ValueError("Sujets.xlsx: colonne 'Numero' introuvable.")
    if idx_titre is None:
        # on tolère titre absent, mais on mettra un fallback
        idx_titre = None

    sujets: List[SujetRef] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        num = _safe_int(row[idx_num] if idx_num < len(row) else None)
        if num is None:
            continue
        titre = ""
        if idx_titre is not None and idx_titre < len(row):
            titre = _norm_str(row[idx_titre])
        if not titre:
            titre = f"Sujet {num}"
        sujets.append(SujetRef(numero=num, titre=titre))

    # tri par numero
    sujets.sort(key=lambda s: s.numero)
    return sujets


def _dedupe_interventions(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Dédoublonnage conservateur: (timecode, auteur, texte) en minuscule/trim.
    On conserve la première occurrence.
    """
    seen: set[Tuple[str, str, str]] = set()
    out: List[Dict[str, Any]] = []
    for it in items or []:
        if not isinstance(it, dict):
            continue
        tc = _norm_str(it.get("timecode")).lower()
        au = _norm_str(it.get("auteur")).lower()
        tx = _norm_str(it.get("texte")).lower()
        key = (tc, au, tx)
        if key in seen:
            continue
        seen.add(key)
        out.append(it)
    return out


def _filter_debrief_for_sujet(debrief_obj: Any, numero: int) -> Dict[str, Any]:
    """
    Extrait uniquement la partie debrief pour un sujet.
    Schéma attendu (provenant de votre passe 1B):
      { "sujets":[{ "numero":int, ... }], "demandes_documents_hors_sujet":[...], "global_debrief":{...} }
    Retourne un objet conforme minimal:
      { "sujet": <obj ou null> }
    """
    res: Dict[str, Any] = {"sujet": None}
    if not isinstance(debrief_obj, dict):
        return res
    sujets = debrief_obj.get("sujets")
    if isinstance(sujets, list):
        for s in sujets:
            if isinstance(s, dict) and _safe_int(s.get("numero")) == numero:
                res["sujet"] = s
                break
    return res


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--global", dest="global_path", required=True, help="Chemin vers global.json (Passe 2A)")
    ap.add_argument("--sujets", dest="sujets_xlsx", required=True, help="Chemin vers Sujets.xlsx")
    ap.add_argument("--out", dest="out_dir", required=True, help="Répertoire de sortie (ex: out/sujets)")
    ap.add_argument("--debrief", dest="debrief_path", default="", help="Optionnel: chemin vers debrief.json")
    ap.add_argument("--include-empty", action="store_true", help="Écrit aussi les sujets sans interventions (défaut: oui)")
    ap.add_argument("--no-dedupe", action="store_true", help="Désactive le dédoublonnage conservateur")
    args = ap.parse_args()

    _ensure_dir(args.out_dir)

    # 1) Référentiel sujets (ordre + titres)
    sujets_ref = _load_sujets_xlsx(args.sujets_xlsx)

    # 2) Global.json
    g = _read_json(args.global_path)
    if not isinstance(g, dict) or "sujets" not in g or not isinstance(g["sujets"], dict):
        raise ValueError("global.json: structure inattendue. Attendu: { 'sujets': { '1': [ ... ], ... } }")

    sujets_map: Dict[str, Any] = g["sujets"]

    # 3) Debrief optionnel
    debrief_obj: Any = None
    if args.debrief_path and os.path.exists(args.debrief_path):
        debrief_obj = _read_json(args.debrief_path)

    # 4) Écriture 1 fichier par sujet du référentiel
    index: List[Dict[str, Any]] = []
    for sref in sujets_ref:
        key = str(sref.numero)
        raw_items = sujets_map.get(key, [])
        items: List[Dict[str, Any]] = raw_items if isinstance(raw_items, list) else []

        if not args.no_dedupe:
            items = _dedupe_interventions(items)

        out_obj: Dict[str, Any] = {
            "sujet": {
                "numero": sref.numero,
                "titre": sref.titre,
            },
            "interventions": items,
        }

        if debrief_obj is not None:
            out_obj["debrief"] = _filter_debrief_for_sujet(debrief_obj, sref.numero)
        else:
            out_obj["debrief"] = {"sujet": None}

        fname = f"sujet_{sref.numero:02d}.json"
        out_path = os.path.join(args.out_dir, fname)

        # Par défaut, on écrit aussi les sujets vides (c'est votre exigence "tous sujets traités")
        if items or args.include_empty:
            _write_json(out_path, out_obj)

        index.append(
            {
                "numero": sref.numero,
                "titre": sref.titre,
                "file": fname,
                "nb_interventions": len(items),
            }
        )

    # 5) Index de contrôle
    idx_path = os.path.join(args.out_dir, "_index.json")
    _write_json(idx_path, {"count": len(index), "items": index})

    print(f"OK: {len(index)} sujets écrits dans: {args.out_dir}")
    print(f"Index: {idx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
